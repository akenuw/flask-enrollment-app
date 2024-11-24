import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# File to save employee data
file_name = "employee_data_with_web_access.xlsx"

# Salary dictionary for predefined categories
salary_dict = {
    "Janitor": 50000,
    "Receptionist": 70000,
    "Cleaner": 50000,
    "Security": 100000,
    "Attendant": 50000
}

# Function to create an Excel file if it doesn't exist
def create_excel_file():
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Employee Data"
        sheet.append(["Name", "Role", "Date and Time of Employment", "Category", "Salary", "Account Number", "Account Name", "Bank Name"])  # Header row
        workbook.save(file_name)

# Function to check for duplicates in the Excel file
def is_duplicate(name, role):
    if not os.path.exists(file_name):
        return False

    workbook = load_workbook(file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] == name and row[1] == role:  # Check if Name and Role match
            return True
    return False

# Function to handle form submission
def submit_form():
    name = name_entry.get()
    role = role_entry.get()
    date_of_employment = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Current date and time
    category = category_var.get()
    salary = salary_entry.get()
    account_number = account_number_entry.get()
    account_name = account_name_entry.get()
    bank_name = bank_name_entry.get()

    if not name or not role or not category or not salary or not account_number or not account_name or not bank_name:
        messagebox.showerror("Error", "All fields are required!")
        return

    # Check for duplicates
    if is_duplicate(name, role):
        messagebox.showerror("Error", "This employee is already enrolled!")
        return

    try:
        # Save data to Excel file
        if not os.path.exists(file_name):
            create_excel_file()

        workbook = load_workbook(file_name)
        sheet = workbook.active
        sheet.append([name, role, date_of_employment, category, salary, account_number, account_name, bank_name])
        workbook.save(file_name)

        messagebox.showinfo("Success", "Employee data saved successfully!")
        clear_form()

    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")

# Function to clear the form
def clear_form():
    name_entry.delete(0, tk.END)
    role_entry.delete(0, tk.END)
    category_var.set(None)
    salary_entry.config(state=tk.NORMAL)
    salary_entry.delete(0, tk.END)
    account_number_entry.delete(0, tk.END)
    account_name_entry.delete(0, tk.END)
    bank_name_entry.delete(0, tk.END)

# Function to update the salary field based on the selected category
def update_salary(*args):
    category = category_var.get()
    if category in salary_dict:
        salary_entry.config(state=tk.NORMAL)  # Temporarily enable to update
        salary_entry.delete(0, tk.END)
        salary_entry.insert(0, salary_dict[category])
        salary_entry.config(state="readonly")  # Lock the field
    elif category == "Others":
        salary_entry.config(state=tk.NORMAL)  # Enable manual input for "Others"
        salary_entry.delete(0, tk.END)

# Create the Excel file if it doesn't exist
create_excel_file()

# Main window setup
root = tk.Tk()
root.title("Employee Enrollment")
root.geometry("600x500")
root.configure(bg="#1e1e2e")

# Title Label
title_label = tk.Label(root, text="Employee Enrollment System", font=("Arial", 16, "bold"), bg="#1e1e2e", fg="white")
title_label.pack(pady=10)

# Frame for input fields
form_frame = tk.Frame(root, bg="#2e2e3e", padx=10, pady=10)
form_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)

# Name
tk.Label(form_frame, text="Name:", bg="#2e2e3e", fg="white").grid(row=0, column=0, sticky="w", pady=5)
name_entry = tk.Entry(form_frame, width=40)
name_entry.grid(row=0, column=1, pady=5)

# Role
tk.Label(form_frame, text="Role:", bg="#2e2e3e", fg="white").grid(row=1, column=0, sticky="w", pady=5)
role_entry = tk.Entry(form_frame, width=40)
role_entry.grid(row=1, column=1, pady=5)

# Category
tk.Label(form_frame, text="Category:", bg="#2e2e3e", fg="white").grid(row=2, column=0, sticky="w", pady=5)
category_var = tk.StringVar()
category_combobox = ttk.Combobox(form_frame, textvariable=category_var, values=["Janitor", "Security", "Receptionist", "Attendant", "Cleaner", "Others"], width=37)
category_combobox.grid(row=2, column=1, pady=5)
category_var.trace("w", update_salary)  # Trigger update_salary when category changes

# Salary
tk.Label(form_frame, text="Salary:", bg="#2e2e3e", fg="white").grid(row=3, column=0, sticky="w", pady=5)
salary_entry = tk.Entry(form_frame, width=40)
salary_entry.grid(row=3, column=1, pady=5)

# Account Number
tk.Label(form_frame, text="Account Number:", bg="#2e2e3e", fg="white").grid(row=4, column=0, sticky="w", pady=5)
account_number_entry = tk.Entry(form_frame, width=40)
account_number_entry.grid(row=4, column=1, pady=5)

# Account Name
tk.Label(form_frame, text="Account Name:", bg="#2e2e3e", fg="white").grid(row=5, column=0, sticky="w", pady=5)
account_name_entry = tk.Entry(form_frame, width=40)
account_name_entry.grid(row=5, column=1, pady=5)

# Bank Name
tk.Label(form_frame, text="Bank Name:", bg="#2e2e3e", fg="white").grid(row=6, column=0, sticky="w", pady=5)
bank_name_entry = tk.Entry(form_frame, width=40)
bank_name_entry.grid(row=6, column=1, pady=5)

# Buttons
button_frame = tk.Frame(root, bg="#1e1e2e")
button_frame.pack(fill=tk.X, pady=10)

clear_button = tk.Button(button_frame, text="Clear", command=clear_form, bg="orange", fg="white", width=15)
clear_button.pack(side=tk.LEFT, padx=20)

submit_button = tk.Button(button_frame, text="Submit", command=submit_form, bg="green", fg="white", width=15)
submit_button.pack(side=tk.RIGHT, padx=20)

# Run the application
root.mainloop()
